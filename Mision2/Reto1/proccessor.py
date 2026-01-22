import re
from openpyxl import load_workbook
# ================================
# FUNCION clean_id
# Elimina caracteres no numéricos de un documento
# "cc75.888.56" = "7588856"
# ================================
def clean_id(value):
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))
# ================================
#  FUNCION merge_name
#  Une nombre y apellido en un solo campo
# ================================
def merge_name(name, lastname):
    if name is None:
        name =""
    if lastname is None:
        lastname =""
    return f"{name} {lastname}".strip()


def process_excel(path):
    #Acceso a la hoja llamada "datos"
    wb= load_workbook(path)
    ws = wb["Datos"]
    #Recorrer todas las filas desde la fila 2
    for row in range(2,ws.max_row+1):
        #Columna D: identificador limpio
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
        )
    
    # Guarde los cambios en el mismo archivo
    wb.save(path)

def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente"
    except PermissionError:
        return(
            False,
            "El archivo Excel está abierto.\n"
            "por Favor, ciérrelo e intente nuevamente."
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False, f"Error inesperado: {str(e)}"
    

from openpyxl import load_workbook


def ejecutar_accion(instruccion, ruta_excel):
    wb = load_workbook(ruta_excel)
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(
                filter(str.isdigit, str(ws[f"{col}{fila}"].value))
            )

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(ruta_excel)
