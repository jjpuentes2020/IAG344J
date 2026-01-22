# processor.py
# LÃ³gica de negocio: operaciones sobre Excel

from openpyxl import load_workbook

def ejecutar_accion(instruccion):
    # Abre el archivo de eje
    # mplo
    wb = load_workbook("ejemplo.xlsx")
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save("ejemplo.xlsx")

def ejecutar_accion(instruccion):
    # Abre el archivo de ejemplo
    wb = load_workbook("ejemplo.xlsx")
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save("ejemplo.xlsx")
